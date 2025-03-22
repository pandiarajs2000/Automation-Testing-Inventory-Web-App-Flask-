import openpyxl
from openpyxl.styles import Font, Color

# read data from the excel
def read_data(file_path, sheet_name, row, column):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    return work_sheet.cell(row, column).value

# row count from the excel
def row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    return work_sheet.max_row

# write data to the excel sheet
def write_data(file_path,sheet_name, row, column,data):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    work_sheet.cell(row,column).value = data
    return workbook.save(file_path)